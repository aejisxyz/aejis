#!/usr/bin/env python3
"""XLSX Processor for Docker container"""

import os
import sys
import json
import base64
from openpyxl import load_workbook

def process_xlsx(file_path: str) -> dict:
    """Process XLSX file and return structured data"""
    try:
        # Read the file
        with open(file_path, 'rb') as f:
            file_data = f.read()
            file_base64 = base64.b64encode(file_data).decode('utf-8')
        
        # Process with openpyxl
        wb = load_workbook(file_path, data_only=True)
        
        # Extract data
        html_parts = []
        html_parts.append('<div class="xlsx-html-content" style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; padding: 20px; background: white;">')
        
        total_rows = 0
        total_cols = 0
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            total_rows += max_row
            total_cols = max(total_cols, max_col)
            
            # Add sheet header
            html_parts.append(f'<h3 style="color: #2563eb; margin: 20px 0 10px 0; border-bottom: 2px solid #e2e8f0; padding-bottom: 5px;">Worksheet: {sheet_name}</h3>')
            
            # Create table
            html_parts.append('<table style="width: 100%; border-collapse: collapse; margin: 10px 0; border: 1px solid #ccc; background: white;">')
            
            # Process rows (limit to first 100 for performance)
            rows_to_process = min(max_row, 100)
            for row_num in range(1, rows_to_process + 1):
                row_data = []
                has_data = False
                
                for col_num in range(1, max_col + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell_value = cell.value
                    
                    if cell_value is not None:
                        cell_text = str(cell_value).strip()
                        if cell_text:
                            has_data = True
                            row_data.append(cell_text)
                        else:
                            row_data.append('')
                    else:
                        row_data.append('')
                
                if has_data:
                    if row_num == 1:  # Header row
                        html_parts.append('<tr style="background: #f8fafc; font-weight: 600;">')
                        for cell_text in row_data:
                            html_parts.append(f'<th style="border: 1px solid #e2e8f0; padding: 8px; text-align: left;">{cell_text}</th>')
                        html_parts.append('</tr>')
                    else:  # Data rows
                        html_parts.append('<tr>')
                        for cell_text in row_data:
                            html_parts.append(f'<td style="border: 1px solid #e2e8f0; padding: 8px;">{cell_text}</td>')
                        html_parts.append('</tr>')
            
            html_parts.append('</table>')
            html_parts.append(f'<p style="color: #64748b; font-size: 0.9rem; margin: 5px 0 20px 0;">📈 {max_row} rows × {max_col} columns</p>')
        
        html_parts.append('</div>')
        html_content = ''.join(html_parts)
        
        # Extract text content
        text_parts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"Sheet: {sheet_name}")
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) for cell in row if cell is not None])
                if row_text.strip():
                    text_parts.append(row_text)
        text_content = '\n'.join(text_parts)
        
        # Extract metadata
        props = wb.properties
        metadata = {
            'title': str(props.title or ''),
            'creator': str(props.creator or ''),
            'subject': str(props.subject or ''),
            'created': str(props.created or ''),
            'modified': str(props.modified or ''),
            'sheets': len(wb.sheetnames),
            'total_rows': total_rows,
            'max_columns': total_cols
        }
        
        return {
            "preview_type": "xlsx",
            "preview_content": text_content or "No content found",
            "xlsx_html": html_content,
            "xlsx_base64": file_base64,
            "base64_preview": file_base64,
            "xlsx_metadata": {
                "sheets": len(wb.sheetnames),
                "total_rows": total_rows,
                "max_columns": total_cols,
                "metadata": metadata
            },
            "success": True
        }
        
    except Exception as e:
        return {
            "preview_type": "xlsx",
            "preview_content": f"XLSX processing error: {str(e)}",
            "success": False,
            "error": str(e)
        }

def main():
    """Main entry point for Docker container"""
    if len(sys.argv) < 2:
        print(json.dumps({"error": "No file path provided", "success": False}))
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    if not os.path.exists(file_path):
        print(json.dumps({"error": "File not found", "success": False}))
        sys.exit(1)
    
    result = process_xlsx(file_path)
    print(json.dumps(result))

if __name__ == "__main__":
    main()

